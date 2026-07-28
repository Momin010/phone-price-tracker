#!/usr/bin/env python3
"""Bundle the deliverable CSVs into one Excel workbook for Art."""
import os, csv, openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(__file__))
D = os.path.join(ROOT, "deliverables")
SHEETS = [
    ("master_sites.csv", "All Sites"),
    ("login_gated.csv", "Login-Gated Shops"),
    ("list_A_original.csv", "A - Glass+Pulled+Refurb"),
    ("list_B_flex_fog.csv", "B - Flex+Fog"),
    ("list_aftermarket.csv", "Aftermarket (OLED-LCD)"),
]
wb = openpyxl.Workbook(); wb.remove(wb.active)
hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1F4E78")
for fn, title in SHEETS:
    p = os.path.join(D, fn)
    if not os.path.exists(p): continue
    ws = wb.create_sheet(title[:31])
    for r, row in enumerate(csv.reader(open(p)), 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if r == 1: cell.font = hdr; cell.fill = fill
    for col in ws.columns:
        w = min(max((len(str(c.value or "")) for c in col), default=10) + 2, 60)
        ws.column_dimensions[col[0].column_letter].width = w
    ws.freeze_panes = "A2"
out = os.path.join(ROOT, "deliverables", "Art_phone_screen_prices.xlsx")
wb.save(out)
print("Wrote", out)
